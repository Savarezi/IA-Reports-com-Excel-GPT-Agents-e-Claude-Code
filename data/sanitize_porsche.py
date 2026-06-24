#!/usr/bin/env python3
"""
Agente de Sanitização de Vendas Porsche
=========================================
Lê uma planilha .xlsx com dados brutos de vendas e aplica as regras de
normalização definidas em `prompet.md` (Porsche Sales Sanitization Schema),
gerando um novo arquivo .xlsx com as colunas sanitizadas inseridas
imediatamente após cada coluna de origem.

Uso:
    python data/sanitize_porsche.py /planilha_base_porsche.xlsx

Se o arquivo de saída não for informado, o agente cria
"<nome_entrada>_sanitized.xlsx" na mesma pasta.
"""

import sys
import re
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ════════════════════════════════════════════════════════════════════════════
# TABELAS DE REFERÊNCIA
# ════════════════════════════════════════════════════════════════════════════

STATE_MAP = {
    'alabama': 'AL', 'alaska': 'AK', 'arizona': 'AZ', 'arkansas': 'AR',
    'california': 'CA', 'colorado': 'CO', 'connecticut': 'CT', 'delaware': 'DE',
    'florida': 'FL', 'georgia': 'GA', 'hawaii': 'HI', 'idaho': 'ID',
    'illinois': 'IL', 'indiana': 'IN', 'iowa': 'IA', 'kansas': 'KS',
    'kentucky': 'KY', 'louisiana': 'LA', 'maine': 'ME', 'maryland': 'MD',
    'massachusetts': 'MA', 'michigan': 'MI', 'minnesota': 'MN',
    'mississippi': 'MS', 'missouri': 'MO', 'montana': 'MT', 'nebraska': 'NE',
    'nevada': 'NV', 'new hampshire': 'NH', 'new jersey': 'NJ',
    'new mexico': 'NM', 'new york': 'NY', 'north carolina': 'NC',
    'north dakota': 'ND', 'ohio': 'OH', 'oklahoma': 'OK', 'oregon': 'OR',
    'pennsylvania': 'PA', 'rhode island': 'RI', 'south carolina': 'SC',
    'south dakota': 'SD', 'tennessee': 'TN', 'texas': 'TX', 'utah': 'UT',
    'vermont': 'VT', 'virginia': 'VA', 'washington': 'WA',
    'west virginia': 'WV', 'wisconsin': 'WI', 'wyoming': 'WY',
    # abreviações (case-insensitive, mapeadas para si mesmas em upper)
    'al': 'AL', 'ak': 'AK', 'az': 'AZ', 'ar': 'AR', 'ca': 'CA', 'co': 'CO',
    'ct': 'CT', 'de': 'DE', 'fl': 'FL', 'ga': 'GA', 'hi': 'HI', 'id': 'ID',
    'il': 'IL', 'in': 'IN', 'ia': 'IA', 'ks': 'KS', 'ky': 'KY', 'la': 'LA',
    'me': 'ME', 'md': 'MD', 'ma': 'MA', 'mi': 'MI', 'mn': 'MN', 'ms': 'MS',
    'mo': 'MO', 'mt': 'MT', 'ne': 'NE', 'nv': 'NV', 'nh': 'NH', 'nj': 'NJ',
    'nm': 'NM', 'ny': 'NY', 'nc': 'NC', 'nd': 'ND', 'oh': 'OH', 'ok': 'OK',
    'or': 'OR', 'pa': 'PA', 'ri': 'RI', 'sc': 'SC', 'sd': 'SD', 'tn': 'TN',
    'tx': 'TX', 'ut': 'UT', 'vt': 'VT', 'va': 'VA', 'wa': 'WA', 'wv': 'WV',
    'wi': 'WI', 'wy': 'WY',
}

WORD_TO_NUM = {
    'zero': 0, 'one': 1, 'two': 2, 'three': 3, 'four': 4, 'five': 5,
    'six': 6, 'seven': 7, 'eight': 8, 'nine': 9, 'ten': 10, 'eleven': 11,
    'twelve': 12, 'thirteen': 13, 'fourteen': 14, 'fifteen': 15,
    'sixteen': 16, 'seventeen': 17, 'eighteen': 18, 'nineteen': 19,
    'twenty': 20, 'thirty': 30, 'forty': 40, 'fifty': 50, 'sixty': 60,
    'seventy': 70, 'eighty': 80, 'ninety': 90, 'hundred': 100,
    'thousand': 1000,
}

CANONICAL_MODELS = [
    '911 Carrera', '911 Carrera S', '911 Carrera GTS', '911 Turbo',
    '911 Turbo S', '911 GT3', '911 GT3 RS', '911 Dakar', '911 Targa 4',
    '911 Targa 4S', '718 Cayman', '718 Cayman S', '718 Cayman GT4 RS',
    '718 Boxster', '718 Boxster GTS', '718 Spyder RS', 'Cayenne',
    'Cayenne S', 'Cayenne Coupe', 'Cayenne E-Hybrid', 'Cayenne Turbo',
    'Cayenne Turbo GT', 'Macan', 'Macan S', 'Macan T', 'Macan GTS',
    'Macan Electric', 'Panamera', 'Panamera 4', 'Panamera 4S',
    'Panamera Turbo', 'Panamera Turbo S', 'Panamera 4 E-Hybrid', 'Taycan',
    'Taycan 4S', 'Taycan GTS', 'Taycan Turbo', 'Taycan Turbo S',
    'Taycan Cross Turismo',
]

PAYMENT_MAP = {
    'credit card': 'Credit Card', 'creditcard': 'Credit Card',
    'credit': 'Credit Card', 'credit card payment': 'Credit Card',
    'debit card': 'Debit Card', 'debit': 'Debit Card',
    'bank transfer': 'Bank Transfer', 'bank_transfer': 'Bank Transfer',
    'bank wire': 'Wire Transfer', 'wire transfer': 'Wire Transfer',
    'wiretransfer': 'Wire Transfer', 'wire': 'Wire Transfer',
    'financing': 'Financing', 'financing plan': 'Financing',
    'finance': 'Financing',
    'lease': 'Lease', 'lease plan': 'Lease', 'leasing': 'Lease',
    'cash': 'Cash', 'cash payment': 'Cash',
    'ach payment': 'ACH Payment', 'ach': 'ACH Payment',
    'crypto payment': 'Crypto Payment', 'crypto': 'Crypto Payment',
}

DELIVERY_MAP = {
    'delivered': 'Delivered', 'deliverd': 'Delivered',
    'in transit': 'In Transit', 'intransit': 'In Transit',
    'pending': 'Pending',
    'cancelled': 'Cancelled', 'canceled': 'Cancelled',
    'awaiting delivery': 'Awaiting Delivery',
    'awaiting pickup': 'Awaiting Pickup',
    'pending approval': 'Pending Approval',
    'pending review': 'Pending Review',
    'shipped': 'Shipped',
    'awaiting review': 'Awaiting Review',
}

MONTH_MAP = {
    'january': '01', 'february': '02', 'march': '03', 'april': '04',
    'may': '05', 'june': '06', 'july': '07', 'august': '08',
    'september': '09', 'october': '10', 'november': '11', 'december': '12',
    'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04', 'jun': '06',
    'jul': '07', 'aug': '08', 'sep': '09', 'oct': '10', 'nov': '11',
    'dec': '12',
}

# Mapeamento: coluna origem -> coluna sanitizada (define ordem de inserção)
SOURCE_TO_SANITIZED = {
    'sale_date':        'SaleDateSanitized',
    'porsche_model':    'PorscheModelSanitized',
    'model_year':       'ModelYearSanitized',
    'sale_price':       'SalesPriceSanitized',
    'vehicle_mileage':  'VehicleMileageSanitized',
    'payment_method':   'PayMethodSanitized',
    'city':             'CitySanitized',
    'state':            'StateSanitized',
    'delivery_status':  'DeliveryStatusSanitized',
}


# ════════════════════════════════════════════════════════════════════════════
# FUNÇÕES AUXILIARES
# ════════════════════════════════════════════════════════════════════════════

def words_to_number(text):
    """Converte números escritos por extenso (inglês) em inteiro."""
    text = text.lower().strip()
    tokens = re.split(r'[\s\-]+', text)
    result, current = 0, 0
    for t in tokens:
        if t not in WORD_TO_NUM:
            return None
        n = WORD_TO_NUM[t]
        if n == 1000:
            current = (current if current else 1) * 1000
            result += current
            current = 0
        elif n == 100:
            current = (current if current else 1) * 100
        else:
            current += n
    return result + current


# ── Datas ─────────────────────────────────────────────────────────────────

def sanitize_date(raw):
    if pd.isna(raw) or str(raw).strip() == '':
        return 'INVALID'
    s = str(raw).strip()
    s = re.sub(r'\s+\d{2}:\d{2}:\d{2}$', '', s).strip()  # remove timestamp

    def valid(y, m, d):
        try:
            datetime(int(y), int(m), int(d))
            return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
        except (ValueError, OverflowError):
            return 'INVALID'

    # YYYY-MM-DD / YYYY/MM/DD / YYYY.MM.DD
    m = re.match(r'^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})$', s)
    if m:
        return valid(m.group(1), m.group(2), m.group(3))

    # MM/DD/YYYY ou MM-DD-YYYY
    m = re.match(r'^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$', s)
    if m:
        return valid(m.group(3), m.group(1), m.group(2))

    # MM/DD/YY ou MM-DD-YY
    m = re.match(r'^(\d{1,2})[/-](\d{1,2})[/-](\d{2})$', s)
    if m:
        return valid(2000 + int(m.group(3)), m.group(1), m.group(2))

    # "Month DDth, YYYY" / "Mon DDth YYYY"
    m = re.match(r'^([A-Za-z]+)\s+(\d{1,2})(?:st|nd|rd|th)?,?\s+(\d{4})$', s, re.I)
    if m:
        mo = MONTH_MAP.get(m.group(1).lower())
        if mo:
            return valid(m.group(3), mo, m.group(2))

    # "DD Month YYYY"
    m = re.match(r'^(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})$', s, re.I)
    if m:
        mo = MONTH_MAP.get(m.group(2).lower())
        if mo:
            return valid(m.group(3), mo, m.group(1))

    return 'INVALID'


# ── Modelo Porsche ────────────────────────────────────────────────────────

def sanitize_model(raw):
    if pd.isna(raw) or str(raw).strip() == '':
        return 'INVALID'
    s_norm = re.sub(r'\s+', ' ', str(raw).strip())
    for canon in CANONICAL_MODELS:
        if s_norm.lower() == canon.lower():
            return canon
    best = None
    for canon in CANONICAL_MODELS:
        if canon.lower() in s_norm.lower() and (best is None or len(canon) > len(best)):
            best = canon
    return best if best else s_norm.title()


# ── Ano do modelo ─────────────────────────────────────────────────────────

def sanitize_year(raw):
    if pd.isna(raw) or str(raw).strip() == '':
        return 'INVALID'
    s = str(raw).strip()

    m = re.match(r'^(\d{4})$', s)
    if m:
        y = int(m.group(1))
        return str(y) if 1990 <= y <= 2035 else 'INVALID'

    # formas compactas "20-24" / "20 24" -> 2024
    m = re.match(r'^(\d{2})[\s\-](\d{2})$', s)
    if m:
        y = int(m.group(1) + m.group(2))
        return str(y) if 1990 <= y <= 2035 else 'INVALID'

    slow = s.lower()
    n = words_to_number(slow)
    if n is not None and 1990 <= n <= 2035:
        return str(n)

    # "twenty twenty four" -> divide em duas metades (20 | 24)
    tokens = re.split(r'[\s\-]+', slow)
    if len(tokens) >= 2:
        for split in range(1, len(tokens)):
            first = words_to_number(' '.join(tokens[:split]))
            second = words_to_number(' '.join(tokens[split:]))
            if first is not None and second is not None:
                if first < 100 and second < 100:
                    y = int(f"{first:02d}{second:02d}")
                else:
                    y = first * 100 + second
                if 1990 <= y <= 2035:
                    return str(y)

    return 'INVALID'


# ── Preço de venda ────────────────────────────────────────────────────────

def sanitize_price(raw):
    if pd.isna(raw) or str(raw).strip() == '':
        return 'INVALID'
    s = str(raw).strip().lower()

    # forma textual: "eighty two thousand USD", "two hundred thousand USD"
    word_part = re.sub(r'usd|dollars?|\$', '', s).strip()
    if re.search(r'[a-z]', word_part.replace('k', '')):
        n = words_to_number(word_part)
        if n is not None:
            return f"{float(n):.2f}"

    s = re.sub(r'\busd\b|\$', '', s)
    s = re.sub(r'\bdollars?\b', '', s).strip()

    # formato europeu "103.750,00" (ponto = milhar, vírgula = decimal)
    if re.search(r'\d+\.\d{3},\d{2}', s):
        s = re.sub(r'\.', '', s).replace(',', '.')
    # formato europeu sem decimal "68.900" (ponto = milhar)
    elif re.match(r'^\d{1,3}(?:\.\d{3})+$', s.strip()):
        s = s.replace('.', '')
    else:
        s = s.replace(',', '')

    m = re.match(r'^([\d.]+)\s*k$', s.strip())
    if m:
        return f"{float(m.group(1)) * 1000:.2f}"

    try:
        return f"{float(s.strip()):.2f}"
    except ValueError:
        return 'INVALID'


# ── Quilometragem ─────────────────────────────────────────────────────────

def sanitize_mileage(raw):
    if pd.isna(raw) or str(raw).strip() == '':
        return 'INVALID'
    s = str(raw).strip().lower()

    if re.search(r'\bnew\b|zero miles?|0 mi', s) or s in ('new', 'new car', 'zero', '0 mi', '0 miles'):
        return 0

    is_km = 'km' in s
    s_clean = re.sub(r'km|miles?|mi\.?|miles?:?', '', s).strip()
    s_clean = re.sub(r'[:,]', '', s_clean).strip()

    if re.search(r'[a-z]', s_clean):
        n = words_to_number(s_clean)
        if n is not None:
            return round(n * 0.621371) if is_km else n
        return 'INVALID'

    # formato europeu "1.200" (ponto = separador de milhar)
    m = re.match(r'^(\d+)\.(\d{3})$', s_clean)
    if m:
        val = int(m.group(1) + m.group(2))
        return round(val * 0.621371) if is_km else val

    s_clean = s_clean.replace(',', '')
    try:
        val = int(float(s_clean))
        return round(val * 0.621371) if is_km else val
    except ValueError:
        return 'INVALID'


# ── Método de pagamento ───────────────────────────────────────────────────

def sanitize_payment(raw):
    if pd.isna(raw) or str(raw).strip() == '':
        return 'INVALID'
    s = re.sub(r'[\-_]', ' ', str(raw).strip().lower())
    s = re.sub(r'\s+', ' ', s).strip()
    if s in PAYMENT_MAP:
        return PAYMENT_MAP[s]
    for k, v in PAYMENT_MAP.items():
        if k in s:
            return v
    return str(raw).strip().title()


# ── Cidade ────────────────────────────────────────────────────────────────

def sanitize_city(raw):
    if pd.isna(raw) or str(raw).strip() == '':
        return 'INVALID'
    return str(raw).strip().title()


# ── Estado ────────────────────────────────────────────────────────────────

def sanitize_state(raw):
    if pd.isna(raw) or str(raw).strip() == '':
        return 'INVALID'
    return STATE_MAP.get(str(raw).strip().lower(), 'INVALID')


# ── Status de entrega ─────────────────────────────────────────────────────

def sanitize_delivery(raw):
    if pd.isna(raw) or str(raw).strip() == '':
        return 'INVALID'
    s = re.sub(r'[!.\s]+$', '', str(raw).strip().lower())
    s = re.sub(r'\s+', ' ', s).strip()
    if s in DELIVERY_MAP:
        return DELIVERY_MAP[s]
    s2 = s.replace('-', ' ').strip()
    if s2 in DELIVERY_MAP:
        return DELIVERY_MAP[s2]
    for k, v in DELIVERY_MAP.items():
        if k in s2:
            return v
    return str(raw).strip().title()


# ════════════════════════════════════════════════════════════════════════════
# PIPELINE PRINCIPAL
# ════════════════════════════════════════════════════════════════════════════

def sanitize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Aplica todas as regras de sanitização e retorna o DataFrame
    com as colunas sanitizadas inseridas imediatamente após cada origem."""

    df = df.copy()
    df['SaleDateSanitized']       = df['sale_date'].apply(sanitize_date)
    df['PorscheModelSanitized']   = df['porsche_model'].apply(sanitize_model)
    df['ModelYearSanitized']      = df['model_year'].apply(sanitize_year)
    df['SalesPriceSanitized']     = df['sale_price'].apply(sanitize_price)
    df['VehicleMileageSanitized'] = df['vehicle_mileage'].apply(sanitize_mileage)
    df['PayMethodSanitized']      = df['payment_method'].apply(sanitize_payment)
    df['CitySanitized']           = df['city'].apply(sanitize_city)
    df['StateSanitized']          = df['state'].apply(sanitize_state)
    df['DeliveryStatusSanitized'] = df['delivery_status'].apply(sanitize_delivery)

    ordered_cols = []
    for col in df.columns:
        if col in SOURCE_TO_SANITIZED.values():
            continue  # será inserida na posição correta abaixo
        ordered_cols.append(col)
        if col in SOURCE_TO_SANITIZED:
            ordered_cols.append(SOURCE_TO_SANITIZED[col])

    return df[ordered_cols]


def apply_formatting(path: str):
    """Aplica formatação visual: cabeçalhos, cores para INVALID, zebra-striping."""
    wb = load_workbook(path)
    ws = wb.active

    header_orig_fill = PatternFill('solid', start_color='1F3864', end_color='1F3864')
    header_san_fill   = PatternFill('solid', start_color='2E75B6', end_color='2E75B6')
    header_font       = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    invalid_fill      = PatternFill('solid', start_color='FFE0E0', end_color='FFE0E0')
    invalid_font      = Font(color='C00000', bold=True, name='Arial', size=9)
    san_fill          = PatternFill('solid', start_color='EBF3FB', end_color='EBF3FB')
    san_font          = Font(color='1F3864', name='Arial', size=9)
    orig_font         = Font(name='Arial', size=9)
    alt_fill          = PatternFill('solid', start_color='F5F5F5', end_color='F5F5F5')
    thin              = Side(style='thin', color='CCCCCC')
    border            = Border(left=thin, right=thin, top=thin, bottom=thin)

    sanitized_headers = set(SOURCE_TO_SANITIZED.values())

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        cell.fill = header_san_fill if cell.value in sanitized_headers else header_orig_fill
    ws.row_dimensions[1].height = 32

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        alt = (row_idx % 2 == 0)
        for cell in row:
            header = ws.cell(1, cell.column).value
            is_san = header in sanitized_headers
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if str(cell.value) == 'INVALID':
                cell.fill = invalid_fill
                cell.font = invalid_font
            elif is_san:
                cell.fill = san_fill
                cell.font = san_font
            else:
                cell.fill = alt_fill if alt else PatternFill()
                cell.font = orig_font

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 30)

    ws.freeze_panes = 'A2'
    wb.save(path)


def run_quality_checks(df: pd.DataFrame) -> dict:
    """Confere as regras de qualidade descritas no schema e retorna um resumo."""
    report = {}
    for san_col in SOURCE_TO_SANITIZED.values():
        report[san_col] = int((df[san_col].astype(str) == 'INVALID').sum())
    report['total_rows'] = len(df)
    report['blank_in_sanitized'] = int(
        df[list(SOURCE_TO_SANITIZED.values())].isna().sum().sum()
    )
    return report


def main():
    if len(sys.argv) < 2:
        print("Uso: python sanitize_porsche.py <entrada.xlsx> [saida.xlsx]")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    if not input_path.exists():
        print(f"Erro: arquivo não encontrado: {input_path}")
        sys.exit(1)

    if len(sys.argv) >= 3:
        output_path = Path(sys.argv[2])
    else:
        output_path = input_path.with_name(input_path.stem + "_sanitized.xlsx")

    print(f"Lendo: {input_path}")
    df_raw = pd.read_excel(input_path, dtype=str)

    required = list(SOURCE_TO_SANITIZED.keys())
    missing = [c for c in required if c not in df_raw.columns]
    if missing:
        print(f"Erro: colunas obrigatórias ausentes no arquivo de entrada: {missing}")
        sys.exit(1)

    print("Aplicando regras de sanitização do schema...")
    df_clean = sanitize_dataframe(df_raw)

    print(f"Salvando: {output_path}")
    df_clean.to_excel(output_path, index=False)
    apply_formatting(str(output_path))

    report = run_quality_checks(df_clean)
    print("\n=== Relatório de Qualidade ===")
    print(f"Total de linhas processadas: {report['total_rows']}")
    print(f"Valores em branco indevidos em colunas sanitizadas: {report['blank_in_sanitized']}")
    for col in SOURCE_TO_SANITIZED.values():
        print(f"  {col}: {report[col]} INVALID")
    print(f"\nArquivo gerado com sucesso: {output_path}")


if __name__ == "__main__":
    main()